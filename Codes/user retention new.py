from matplotlib import pyplot as plt
import datetime
import time
import openpyxl
import json
from openpyxl import Workbook

start = "2020-11-01T15:00:00.000" #Retention 계산을 시작할 주의 일요일로 설정 : 날짜가 2020-12-03이면 2020-12-02T15:00:00.000000 으로 설정해야함
start = datetime.datetime.strptime(start,'%Y-%m-%dT%H:%M:%S.%f')

excel_date = ["before"] #엑셀에 각 주의 시작날과 같이 저장하기 위함
user_id = [[]] #retention에 사용할 user_name 기록할 리스트, 0번째 인덱스는 start의 이전에 가입한 사람들의 user_name

#회원들의 가입한 날짜가 저장되어있는 엑셀파일을 읽음
username_exl = openpyxl.load_workbook('auth_user_re.xlsx') 
username_exl = username_exl['auth_user_re']
username_exl = username_exl['A':'E']

exl_data = []

excel_start = datetime.datetime(2020,11,1) #엑셀에서 읽기 시작 할 날짜 (가입한 날짜)

#1주 단위로 엑셀 데이터를 읽어서 해당 주에 가입한 사람들의 user_name을 user_id에 넣음(각 주의 시작이 됨)
for cell1,cell2 in zip(username_exl[1],username_exl[4]):
    v = cell2.value
    if (type(cell2.value) == datetime.datetime):
        if (cell2.value>= excel_start and cell2.value< excel_start + datetime.timedelta(days=7)):
            exl_data.append(cell1.value)
        elif (cell2.value>=excel_start + datetime.timedelta(days=7) and cell2.value<excel_start + datetime.timedelta(days=14) ):
            user_id.append(exl_data)
            exl_data = [cell1.value]
            excel_start += datetime.timedelta(days=7)
    else:
        print("error")

user_id.append(exl_data)

result = [] #retention의 최종 결과를 저장하게 될 리스트

#result에 각 주의 시작점이 될 그 주에 가입한 사람들의 명 수를 넣어줌
for x in user_id:
    result.append([len(x)])

now_index = 1

while(start + datetime.timedelta(days=7)<datetime.datetime.now()):

    print(start)
    #extract data from elastic을 통해 json형식으로 저장한 로그 데이터들을 1주 단위로 읽어오기
    json_data = []
    excel_date.append(start.strftime("%Y-%m-%dT%H:%M:%S.%f")[:10])
    for i in range(7):
        start = start.strftime("%Y-%m-%dT%H:%M:%S.%f")
        file_name = start[:10] #파일 이름을 년-월-일.json으로 저장하기 위함
        with open('./log json/'+file_name+'.json', 'r', encoding='utf-8') as read_file:
            datas = json.load(read_file)
        json_data.extend(datas)
        start = datetime.datetime.strptime(start,'%Y-%m-%dT%H:%M:%S.%f')
        start += datetime.timedelta(days=1)
        
    this_week_users = [] #해당 주에 김박사넷에 접속한 username들

    #읽어온 json_data에 username이 있고, 그 username이 아직 this_week_users에 들어가 있지 않은 애들만 넣어줌 -> 중복을 방지함
    for users in json_data:
        if 'username' in users and users["username"] not in this_week_users:
            this_week_users.append(users["username"])

    #현재 계산 중인 주에 가입한 회원들은 당연히 있을거니까, 그 회원들은 제외함
    this_week_users = [x for x in this_week_users if x not in user_id[now_index]]


    for j in range(1,now_index,1):
        not_in_users = [x for x in this_week_users if x not in user_id[j]] #user_id[j], 즉, j번째 주에 가입하지 않은 유저들만 뽑음
        result[j].append(len(this_week_users)-len(not_in_users)) #this_week_users에서 j번째 주에 가입하지 않은 유저들을 빼줌 = this_week_users에서 j번째 주에 가입한 유저들의 수
        this_week_users = not_in_users #this_week_users를 not_in_users로 바꿔준다
    
    result[0].append(len(this_week_users)) #for문을 다 돌리고도 this_weeks_users에 남아있는 username들은 우리가 갖고있는 데이터 이전에 가입한 사람들이므로, 그 사람들의 명 수는 0번째에 넣어준다

    #user_id 0번째 인덱스에 우리가 갖고있는 데이터 이전에 가입한 username이 들어가있는데, 이번 주에 this_week_users 중 지금까지 남아있는 username들과 합치고 중복 제거해서 추가하기 위함
    prep_user = user_id[0]
    prep_user.extend(this_week_users)
    prep_user = set(prep_user)
    user_id[0] = list(prep_user)

    now_index+=1

result[0][0] = len(user_id[0]) #우리가 가지고 있는 데이터 이 전에 가입한 사람들의 총 명수를 입력함(시작값이 됨)


#엑셀로 저장
write_wb = Workbook()

write_ws = write_wb.active
write_ws = write_wb.create_sheet('TEST')

for i in range(len(excel_date)):
    excel_data = []
    excel_data.append(excel_date[i])
    excel_data.extend(result[i])
    write_ws.append(excel_data)

write_wb.save('./User Retention.xlsx')

#retention 그래프 생성
index = [x+1 for x in range(len(result[1]))]

for i in range(1,now_index,1):
    m = result[i][0]
    for x in range(len(result[i])):
        result[i][x] = result[i][x]/m
    plt.plot(index, result[i])
    index.pop()

plt.show()

#0번 인덱스, 우리가 알고있는 데이터 이 전에 가입한 유저들에 대한 그래프
result[0][0] = len(user_id[0])
index = [x+1 for x in range(len(result[0]))]
plt.plot(index,result[0])
plt.show()