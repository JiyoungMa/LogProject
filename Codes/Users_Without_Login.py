import json
from matplotlib import pyplot as plt
import datetime
import openpyxl
from openpyxl import Workbook


start = "2020-11-1T15:00:00.000" #Retention 계산을 시작할 주의 일요일로 설정 : 날짜가 2020-12-03이면 2020-12-02T15:00:00.000000 으로 설정해야함
start = datetime.datetime.strptime(start,'%Y-%m-%dT%H:%M:%S.%f')

user_id = [[]]  #retention에 사용할 user_name 기록할 리스트, 0번째 인덱스는 start의 이전에 가입한 사람들의 user_name

excel_date = ["before"] #엑셀에 각 주의 시작날과 같이 저장하기 위함

#username을 많이 가지고 있는 특이 케이스인 ip들에 대한 정보를 읽어와서 long_ip에 저장하기
longip_exl = openpyxl.load_workbook('long ip.xlsx')
longip_exl = longip_exl['시트1']
longip_exl = longip_exl['A']

long_ip = []

for cell in longip_exl:
    long_ip.append(cell.value)

#회원들의 가입한 날짜가 저장되어있는 엑셀파일을 읽음
username_exl = openpyxl.load_workbook('auth_user_re.xlsx')
username_exl = username_exl['auth_user_re']
username_exl = username_exl['A':'E']

exl_data = []

excel_start = datetime.datetime(2020,11,1)  #엑셀에서 읽기 시작 할 날짜 (가입한 날짜)

#1주 단위로 엑셀 데이터를 읽어서 해당 주에 가입한 사람들의 user_name을 user_id에 넣음(각 주의 시작이 됨)
for cell1,cell2 in zip(username_exl[1],username_exl[4]):
    v = cell2.value
    if (type(cell2.value) == datetime.datetime):
        if (cell2.value>= excel_start and cell2.value< excel_start + datetime.timedelta(days=7)):
            exl_data.append(cell1.value)
        elif (cell2.value>=excel_start + datetime.timedelta(days=7) and cell2.value<excel_start + datetime.timedelta(days=14) ):
            user_id.append(exl_data)
            exl_data = [cell1.value]
            excel_start += datetime.timedelta(days=7)
    else:
        print("error")

user_id.append(exl_data)

user_ip = [] #각 ip가 어떤 username에 속하는지 체크하기 위함

result = [[]] #retention의 최종 결과를 저장하게 될 리스트
now_index = 1

for i in range(len(user_id)):
    user_ip.append(dict()) #ip가 key고 username이 value

while(start + datetime.timedelta(days=7)<datetime.datetime.now()):
    print(start)
    excel_date.append(start.strftime("%Y-%m-%dT%H:%M:%S.%f")[:10])

    #extract data from elastic을 통해 json형식으로 저장한 로그 데이터들을 1주 단위로 읽어오기
    json_data = []
    for i in range(7):
        start = start.strftime("%Y-%m-%dT%H:%M:%S.%f")
        file_name = start[:10]
        with open('./'+file_name+'.json', 'r', encoding='utf-8') as read_file:
            datas = json.load(read_file)
        json_data.extend(datas)
        start = datetime.datetime.strptime(start,'%Y-%m-%dT%H:%M:%S.%f')
        start += datetime.timedelta(days=1)

    today_ip = [] #해당 주에 접속한 username이 없는 ip들

    for users in json_data:
        if 'username' in users: #username이 있는 경우
            ip = users['ip']
            username = users['username']
            b = False
            
            for i in range(0, now_index+1, 1):
                #username이 user_id의 i번째 인덱스에 속해있고(i번째 주에 가입한 user임)
                #ip가 long_ip에 속하지 않고, 아직 user_ip[i]의 key에 있지 않은 경우, key로 추가하고 value로 username를 준다
                if username in user_id[i] and ip not in user_ip[i].keys() and ip not in long_ip: 
                    user_ip[i][ip] = username
                    b = True
                    break
                #username이 user_id의 i번째 인덱스에 속해있고(i번째 주에 가입한 user임)
                #ip가 이미 user_ip[i]의 key에 있는 경우
                elif username in user_id[i] and ip in user_ip[i].keys():
                    b = True
                    break
            
            #username이 user_id에 아예 속해있지 않고, ip가 user_ip[0]의 key로 존재하지 않은 경우
            #갖고 있는 데이터 이 전에 가입한 것으로 판단하고, user_ip[0]에 ip를 key로, username을 value로 추가함
            if b == False and ip not in user_ip[0].keys() and ip not in long_ip:
                user_ip[0][ip] = username
                    
        else: #username이 없는 경우
            ip = users['ip']
            if ip not in today_ip and ip not in long_ip: #long_ip에 속하지 않으면 today_ip에 중복하는지 확인하고 추가함
                today_ip.append(ip)
    
    result.append([])
    for i in range(0,now_index,1):
        registered_user = [] #비로그인인데 가입한 ip인 경우를 넣기 위함
        
        #user_ip의 key 중에 ip가 있다면, 가입유저가 비로그인으로 접속 한 것이라 registered user에 넣어줌
        for ip in today_ip:
            if ip in user_ip[i].keys():
                registered_user.append(user_ip[i][ip])
        registered_user = set(registered_user) #중복제거
        result[i].append(len(registered_user))

    now_index += 1

for l in range(len(result)):
    if result[l] != [] :
        logged_user = user_ip[l].values()
        logged_user = set(logged_user)
        result[l].insert(0,len(logged_user))
        index = [x+1 for x in range(len(result[l]))]
        total = result[0]
        for i in range(len(result[l])):
            result[i] = result[i] / total
        plt.plot(index,result[l])

plt.show()

write_wb=Workbook()

write_ws = write_wb.active
write_ws = write_wb.create_sheet('ip 기준')

for i in range(len(excel_date)):
    excel_data = []
    excel_data.append(excel_date[i])
    excel_data.extend(result[i])
    write_ws.append(excel_data)

write_wb.save('./User Retention.xlsx')
